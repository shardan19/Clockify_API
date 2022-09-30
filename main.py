import jsbeautifier
import requests

import json
import datetime
import isodate
import pytz
import string

from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter

from types import SimpleNamespace
def zaoczasfloat(seconds, zao):
    hou = seconds // 3600
    min = (seconds // 60) % 60
    sec=seconds % 60
    time=0
    time += hou
    if zao == 1:
        if min > 0:
            time += 1
    elif zao == 0.5:
        if (min < 30 and min != 0) or(min == 30 and sec== 0)or(min==0 and sec>0):
            time += 0.5
        elif min != 0:
            time += 1
    elif zao == 0.25:
        if (min < 15 and min != 0 ) or(min <= 15 and min != 0 and sec==0):
            time += 0.25
        elif (min < 30 and min != 0)or(min == 30 and sec== 0):
            time += 0.5
        elif (min <= 45 and min != 0)or(min == 45 and sec== 0):
            time += 0.75
        elif min != 0:
            time += 1
    return time
def zaoczasstring(seconds, zao):
    hou = seconds // 3600
    min = (seconds // 60) % 60
    sec = seconds % 60
    time=0
    times=""


    if zao == 1:
        if min > 0:
            hou += 1
    elif zao == 0.5:
        if (min < 30 and min != 0) or(min == 30 and sec== 0)or(min==0 and sec>0):
            time += 0.5
        elif min != 0:
            hou += 1
    elif zao == 0.25:
        if (min < 15 and min != 0 ) or(min <= 15 and min != 0 and sec==0):
            time += 0.25
        elif (min < 30 and min != 0)or(min == 30 and sec== 0):
            time += 0.5
        elif (min <= 45 and min != 0)or(min == 45 and sec== 0):
            time += 0.75
        elif min != 0:
            hou += 1
    if hou ==1:
        times = str(hou) + " godzina"
    elif hou>1:
        times = str(hou) + " godziny"
    elif hou > 4:
        times = str(hou) + " godzin"
    if time==0.5:
        times=str(times)+" 30 minut"
    elif time == 0.25:
        times = str(times)+" 15 minut"
    elif time ==0.75:
        times =str(times)+ " 45 minut"
    return times
# czas ktory okresla pelny poprzedni miesiac, z tego czasu zostana pobrane taski uzytkownikow
dt = datetime.datetime.today()
if dt.month == 1:
    StartDate = str(dt.year - 1) + "-12-01T00:00:00.998Z"
elif dt.month > 9:
    StartDate = str(dt.year) + "-" + str(dt.month - 1) + "-01T00:00:00.998Z"
else:
    StartDate = str(dt.year) + "-0" + str(dt.month - 1) + "-01T00:00:00.998Z"
EndDate = str(dt.year) + "-0" + str(dt.month) + "-01T00:00:00.998Z"
print(StartDate + "--------------" + EndDate)
# load config
projWithChild, allUsrTasks = [], []
projtasks = []
config = open("config.json", "r")
configcontent = config.read()
configtoobjects = json.loads(configcontent)
# print(configtoobjects['Projekty'])
# ClocifyApi
ApiKey = configtoobjects['ApiKey']
StartDate = configtoobjects['StartDate']
EndDate = configtoobjects['EndDate']
configProjects = configtoobjects['Projekty']
bilable=int(configtoobjects["billable"])
if bilable ==1:
    bilabl="true"
else:
    bilabl="false"
print(StartDate + "--------------" + EndDate)

headers = {'X-Api-Key': ApiKey}
pa = {'content-type': 'application/json', 'page-size': '200'}
projpa = {'content-type': 'application/json', 'page-size': '200', 'billable':bilabl}


client_requesst = requests.get('https://api.clockify.me/api/v1/workspaces/', headers=headers, params=pa)
res_client = jsbeautifier.beautify(client_requesst.text)
res_client_to_objects = json.loads(res_client)

workspaceId = res_client_to_objects[0]['id']
# pobranie wszystkich projektow i ich podprojektow pasujacych do tych z konfiga
url_projects = 'https://api.clockify.me/api/v1/workspaces/' + workspaceId + '/projects/'
projects_request = requests.get(url_projects, headers=headers, params=projpa)
projects_request_clean = jsbeautifier.beautify(projects_request.text)
projects_request_toobject = json.loads(projects_request_clean)

projcounter = 0
for x in projects_request_toobject:
    # print(x['name'])
    projcounter += 1
print("liczba projektow:" + str(projcounter))
url_client = 'https://api.clockify.me/api/v1/workspaces/' + workspaceId + '/clients/'
client_request = requests.get(url_client, headers=headers, params=pa)
client_request_clean = jsbeautifier.beautify(client_request.text)
client_request_toobject = json.loads(client_request_clean)

for project in configProjects:
    clientName = (project['nazwa'])
    client_object_from_api_response = list(filter(lambda x: x["name"] == clientName, client_request_toobject))

    # print(client_object_from_api_response)
    client_id = client_object_from_api_response[0]['id']
    project_object_from_api_response = list(filter(lambda x: x["clientId"] == client_id, projects_request_toobject))
    for proj in project_object_from_api_response:
        proj["configclientname"] = clientName
        # print(proj)
    # print(project_object_from_api_response[0])
    # print(project_object_from_api_response[1])

    projWithChild.append(project_object_from_api_response)
    # print('+')

# wszyscy uzytkownicy
url_users = 'https://api.clockify.me/api/v1/workspaces/' + workspaceId + '/users'
userpa = {'content-type': 'application/json', 'page-size': '400'}
user_request = requests.get(url_users, headers=headers, params=userpa)
user_request_clean = jsbeautifier.beautify(user_request.text)
user_request_to_object = json.loads(user_request_clean)
# print(user_request_clean)

# wszystkie taski uzytkownikow
taskpa = {'content-type': 'application/json', 'page-size': '1000', 'start': StartDate, 'end': EndDate }
wszystkiezadania = 0
for user in user_request_to_object:
    url_user_tasks_times = 'https://api.clockify.me/api/v1/workspaces/' + workspaceId + '/user/' + user[
        'id'] + '/time-entries/'
    task_request = requests.get(url_user_tasks_times, headers=headers, params=taskpa)
    task_request_clean = jsbeautifier.beautify(task_request.text)
    if len(task_request_clean) > 2:

        task_request_to_object = json.loads(task_request.text)
        #print(task_request_to_object)
        # print("-----------------------------")
        for tsku in task_request_to_object:
            #print(tsku)
            wszystkiezadania += 1
            allUsrTasks.append(tsku)

print("wszystkie zadania które otrzymano:" + str(wszystkiezadania))
print(projWithChild)
print(allUsrTasks)
# print("--------------------------------")
# Przyrownanie taskow pasujacych do projektow z configa
counter = 0
print("--------------------------------")
print("allusertaskscount" + str(len(allUsrTasks)))
secotaskcount = 0
tricounetr = 0

for projgroup in projWithChild:
    for project in projgroup:
        project_tasks=[]
        #print(project)
        for task in allUsrTasks:

            if task["projectId"]==project["id"]:
                tricounetr+=1
                project_tasks.append(task)
        #project_tasks = list(filter(lambda x: x["projectId"] == project["id"], allUsrTasks))
        #print(str(len(project_tasks))+"--"+str(project_tasks))
        if len(project_tasks) > 0:

            for el in project_tasks:
                el["configclientname"] = project["configclientname"]
                # print(str(el) + "---" + project["name"])
                secotaskcount += 1
            # project_tasks["cofigclientname"]=project["cofigclientname"]
            projtasks.append(project_tasks)
            # print(project_tasks);
print("drugi licznik zadań==" + str(secotaskcount))
print("trzeci licznik zadań==" + str(tricounetr))

print("--------------------------------")



now_utc=datetime.datetime.utcnow()   #utcnow class method
print(now_utc)
print("strefa czasowa:"+str(now_utc))
pacific = pytz.timezone('US/Pacific')
d = datetime.datetime.now(pacific)
print (d)
# print(projtasks)
# kombajn mielonka
for key, project in enumerate(configProjects):

    xls_proj_first_row = []
    clientName = project['nazwa']
    abologic = int(project['abologic'])
    cenaabo = float(project['cenaabo'])

    xls_proj_first_row.append(clientName)
    xls_proj_first_row.append("Abonament")

    godziny_abo = int(project['godziny_abo'])
    czasy_stawki = project['czasy_stawki']
    weekend = project['weekend']

    przeskoklogicbeta=int(project["przeskoklogicbeta"])
    #print(project['weekend']["zaokraglanie"])

    tab_godz = {}
    abo_godz = {}
    week_godz = {}

    pathxls = clientName+".xlsx"
    wb = Workbook()
    ws = wb.active

    for h_key, conftime in enumerate(czasy_stawki):
        xls_proj_first_row.append("czas:" + conftime["start"] + "-" + conftime["koniec"])
        tab_godz[h_key] = 0
    tab_godz["Sobota"]=0
    tab_godz["Niedziela"]=0

    xls_proj_first_row.append("Sobota")
    xls_proj_first_row.append("niedziela")

    ws.append(xls_proj_first_row)
    print("taskie dla projektu:" + clientName)

    xls_rest_row = []
    xls_rest_row2 = []

    tsccounter = 0
    for tasksinproj in projtasks:
        h_to_xls = 0
        list_of_tasks = list(filter(lambda x: x["configclientname"] == clientName, tasksinproj))
        #print(str(len(list_of_tasks))+"--"+str(list_of_tasks))
        if len(list_of_tasks) > 0:
            for task in list_of_tasks:
                # print(task)
                tsccounter += 1
                # print("nazwa:" + task["description"] + "początek tasku:" + task["timeInterval"][
                #     "start"] + "--koniec:" + task["timeInterval"]["end"] + "--czas trwania:" +
                #       task["timeInterval"]["duration"])
                #print(task["timeInterval"]["start"])
                start = isodate.parse_datetime(task["timeInterval"]["start"])
                start.replace(tzinfo=pytz.UTC)
                start=start.astimezone(pytz.timezone("Europe/Warsaw"))
                print(str(start)+ task["description"]+str(start.weekday()))
                end = isodate.parse_datetime(task["timeInterval"]["end"])
                end.replace(tzinfo=pytz.UTC)
                end = end.astimezone(pytz.timezone("Europe/Warsaw"))


                duration = isodate.parse_duration(task["timeInterval"]["duration"])
                mi = (duration.seconds // 60) % 60
                se = duration.seconds % 60

                # print(start.day)
                # print(end)
                # print(duration)
                taskrow = []
                taskrow.append(task["description"])
                previous = next_ = None
                l = len(czasy_stawki)
                taskrow.append(" ")
                #print(start.weekday())
                for x_key, conftime in enumerate(czasy_stawki):
                    cofigkoniec = (int(conftime["koniec"]))
                    cofigstart = (int(conftime["start"]))
                    czy_abo = int(conftime["abo"])
                    zao = float(conftime["zaokraglanie"])

                    if start.weekday() == 5 and end.weekday() == 6:
                        taskrow.append("przeskoksobonied")
                        tsk_tim_to_add = 0
                        tsk_tim_to_add2 = 0
                        zao1 = float(weekend["zaokraglanieSobo"])
                        zao2 = float(weekend["zaokraglanieNied"])


                        firtime = 24 * 3600 - start.hour * 3600 - start.minute * 60 - start.second

                        czasst = zaoczasstring(firtime, zao1)
                        tsk_tim_to_add = zaoczasfloat(firtime, zao1)
                        zaoxsec = int(zao1 * 3600)
                        taskrow.append("")
                        taskrow.append("")
                        taskrow.append("")
                        taskrow.append("")
                        if przeskoklogicbeta == 1:
                            if zaoxsec < duration.seconds:

                                taskrow[l + 2] = czasst

                                if godziny_abo > 0 and czy_abo == 1 and abologic == 1:
                                    if godziny_abo - tsk_tim_to_add >= 0:
                                        godziny_abo -= tsk_tim_to_add
                                    else:
                                        tocount = tsk_tim_to_add - godziny_abo
                                        tab_godz["Sobota"] += tocount
                                        godziny_abo = 0
                                else:
                                    tab_godz["Sobota"] += tsk_tim_to_add
                            else:
                                taskrow.append(" ")
                        else:
                            taskrow[l + 2] = czasst
                            if godziny_abo > 0 and czy_abo == 1 and abologic == 1:
                                if godziny_abo - tsk_tim_to_add >= 0:
                                    godziny_abo -= tsk_tim_to_add
                                else:
                                    tocount = tsk_tim_to_add - godziny_abo
                                    tab_godz["Sobota"] += tocount
                                    godziny_abo = 0
                            else:
                                tab_godz["Sobota"] += tsk_tim_to_add
                        sectime = duration.seconds - firtime
                        czasst = zaoczasstring(sectime, zao2)
                        taskrow[l + 3] = czasst
                        tsk_tim_to_add2 = zaoczasfloat(sectime, zao2)

                        if godziny_abo > 0 and czy_abo == 1 and abologic == 1:
                            if godziny_abo - tsk_tim_to_add2 >= 0:
                                godziny_abo -= tsk_tim_to_add2
                            else:
                                tocount = tsk_tim_to_add2 - godziny_abo
                                tab_godz["Niedziela"] += tocount
                                godziny_abo = 0
                        else:
                            tab_godz["Niedziela"] += tsk_tim_to_add2
                        break

                    elif(start.weekday()==6 ):
                        print("NIEDZ")
                        taskrow.append("NIED")
                        zao = float(weekend["zaokraglanieNied"])
                        czas = zaoczasfloat(duration.seconds, zao)
                        czasstring = zaoczasstring(duration.seconds, zao)
                        taskrow.append("")
                        taskrow.append("")
                        taskrow.append("")
                        taskrow.append("")
                        taskrow.append("")
                        taskrow.append("")
                        taskrow.append("")
                        taskrow.append("")
                        taskrow[l + 3] = czasstring

                        if godziny_abo > 0 and czy_abo == 1 and abologic == 1:
                            if godziny_abo - czas >= 0:
                                godziny_abo -= czas
                            else:
                                tocount = czas - godziny_abo
                                tab_godz["Niedziela"]+= tocount
                                godziny_abo = 0
                        else:
                            tab_godz["Niedziela"] += czas
                        break
                    elif start.weekday()==5:
                        print("SOBO")
                        taskrow.append("SOBO")
                        zao = float(weekend["zaokraglanieSobo"])
                        czas = zaoczasfloat(duration.seconds, zao)
                        czasstring = zaoczasstring(duration.seconds, zao)
                        taskrow.append("")
                        taskrow.append("")

                        taskrow.append("")
                        taskrow.append("")

                        taskrow[l + 2] = czasstring
                        if godziny_abo > 0 and czy_abo == 1 and abologic == 1:
                            if godziny_abo - czas >= 0:
                                godziny_abo -= czas
                            else:
                                tocount = czas - godziny_abo
                                tab_godz["Sobota"] += tocount
                                godziny_abo = 0
                        else:
                            tab_godz["Sobota"] += czas
                        break
                    elif (start.weekday()==6 or start.weekday()==5)and end.weekday()<5:
                        print("chory przypadek")

                    elif cofigstart <= start.hour and (cofigkoniec > end.hour or(cofigkoniec == end.hour and mi==0 and se==0) ):
                        #print(str(duration) + "od:" + str(start) + "do:" + str(end) + task["description"] + "----")

                        tsk_tim_to_add = 0
                        if end.day != start.day:
                            if x_key+1 ==l:

                                print("przeskok północ ")
                                tsk_tim_to_add = zaoczasfloat(duration.seconds, zao)
                                czasstring = zaoczasstring(duration.seconds, zao)
                                taskrow.append(czasstring)

                                if godziny_abo > 0 and czy_abo == 1 and abologic == 1:
                                    if godziny_abo - tsk_tim_to_add >= 0:
                                        godziny_abo -= tsk_tim_to_add
                                    else:
                                        tocount = tsk_tim_to_add - godziny_abo
                                        tab_godz[x_key] += tocount
                                        godziny_abo = 0
                                else:
                                    tab_godz[x_key] += tsk_tim_to_add
                                taskrow.append("")
                            else:
                                taskrow.append("")


                        else:
                            #print(str(duration) + "od:" + str(start) + "do:" + str(end) + task["description"] + "----")
                                #print("h:" + str(duration.seconds // 3600) + "m:" + str((duration.seconds // 60) % 60) + "s:" + str(duration.seconds % 60))

                            # print(conftime["start"] + "--" + str(start.hour) + "----" + conftime["koniec"] + "--" + str(end.hour))
                            # print("h:" + str(duration.seconds // 3600) + "m:" + str((duration.seconds // 60) % 60) + "s:" + str(duration.seconds % 60))
                            hou = duration.seconds // 3600
                            min = (duration.seconds // 60) % 60

                            zao = float(conftime["zaokraglanie"])
                            czasst=zaoczasstring(duration.seconds,zao)
                            taskrow.append(czasst)

                            tsk_tim_to_add =zaoczasfloat(duration.seconds,zao)

                            if godziny_abo > 0 and czy_abo == 1 and abologic == 1:
                                if godziny_abo - tsk_tim_to_add >= 0:
                                    godziny_abo -= tsk_tim_to_add
                                else:
                                    tocount = tsk_tim_to_add - godziny_abo
                                    tab_godz[x_key] += tocount
                                    godziny_abo = 0
                            else:
                                tab_godz[x_key] += tsk_tim_to_add
                            break
                    elif  end.day != start.day:
                        # print(str(duration)+"od:"+str(start)+"do:"+str(end)+task["description"]+"----")
                        if task["description"] == "Aktualizacja VMware nowego i starego środowiska, aktualizacja i backup BI":
                            print("xxxxxxxxxxxxxxxxxxxxx")
                            print(str(duration) + "od:" + str(start) + "do:" + str(end) + task["description"] + "----")
                            print(zaoxsec)
                            print(duration.seconds)
                        print("evenement")
                        print(task["description"])
                        tsk_tim_to_add = 0
                        # print(conftime["start"] + "--" + str(start.hour) + "----" + conftime["koniec"] + "--" + str(end.hour))
                        # print("h:" + str(duration.seconds // 3600) + "m:" + str((duration.seconds // 60) % 60) + "s:" + str(duration.seconds % 60))
                        hou = duration.seconds // 3600
                        min = (duration.seconds // 60) % 60
                        zao = float(conftime["zaokraglanie"])
                        # print(zao)
                        tsk_tim_to_add = zaoczasfloat(duration.seconds, zao)
                        czasstring = zaoczasstring(duration.seconds, zao)
                        taskrow.append(czasstring  )
                        taskrow.append("SPRAWDZIC"  )

                        if godziny_abo > 0 and czy_abo == 1 and abologic == 1:
                            if godziny_abo - tsk_tim_to_add >= 0:
                                godziny_abo -= tsk_tim_to_add
                            else:
                                tocount = tsk_tim_to_add - godziny_abo
                                tab_godz[x_key] += tocount
                                godziny_abo = 0
                        else:
                            tab_godz[x_key] += tsk_tim_to_add
                    elif cofigstart <= start.hour and start.hour < cofigkoniec and cofigkoniec <= end.hour:
                        # print(str(duration)+"od:"+str(start)+"do:"+str(end)+task["description"]+"----")
                        print("przeskok")
                        #if task["description"] == "Qnap - nocna akcja serwisowa.":
                            #print("tuuuuuuuuuu")
                            #print(str(duration) + "od:" + str(start) + "do:" + str(end) + task["description"] + "----")
                        tsk_tim_to_add = 0
                        tsk_tim_to_add2 = 0

                        firtime = cofigkoniec * 3600 - start.hour * 3600 - start.minute * 60 - start.second

                        czasst = zaoczasstring(firtime, zao)
                        tsk_tim_to_add = zaoczasfloat(firtime, zao)
                        zaoxsec= int(zao* 3600)

                        if przeskoklogicbeta == 1:
                            if zaoxsec < duration.seconds:

                                taskrow.append(czasst)

                                if godziny_abo > 0 and czy_abo == 1 and abologic == 1:
                                    if godziny_abo - tsk_tim_to_add >= 0:
                                        godziny_abo -= tsk_tim_to_add
                                    else:
                                        tocount = tsk_tim_to_add - godziny_abo
                                        tab_godz[x_key] += tocount
                                        godziny_abo = 0
                                else:
                                    tab_godz[x_key] += tsk_tim_to_add
                                sectime = duration.seconds - firtime
                            else:
                                taskrow.append(" ")

                                tab_godz[x_key]=0
                        else:

                            taskrow.append(czasst)

                            if godziny_abo > 0 and czy_abo == 1 and abologic == 1:
                                if godziny_abo - tsk_tim_to_add >= 0:
                                    godziny_abo -= tsk_tim_to_add
                                else:
                                    tocount = tsk_tim_to_add - godziny_abo
                                    tab_godz[x_key] += tocount
                                    godziny_abo = 0
                            else:
                                tab_godz[x_key] += tsk_tim_to_add
                        sectime = duration.seconds - firtime

                        if x_key < l:
                            nextzao = float(czasy_stawki[x_key + 1]["zaokraglanie"])

                            czasst = zaoczasstring(sectime, nextzao)
                            taskrow.append(czasst)
                            tsk_tim_to_add2 = zaoczasfloat(sectime, zao)

                            if godziny_abo > 0 and czy_abo == 1 and abologic == 1:
                                if godziny_abo - tsk_tim_to_add2 >= 0:
                                    godziny_abo -= tsk_tim_to_add2
                                else:
                                    tocount = tsk_tim_to_add2 - godziny_abo
                                    tab_godz[x_key + 1] += tocount
                                    godziny_abo = 0
                            else:
                                tab_godz[x_key + 1] += tsk_tim_to_add2
                        else:
                            nextzao = float(czasy_stawki[0]["zaokraglanie"])
                            czasst = zaoczasstring(sectime, nextzao)
                            taskrow.append(czasst)
                            tsk_tim_to_add2 = zaoczasfloat(sectime,nextzao)

                            if godziny_abo > 0 and czy_abo == 1 and abologic == 1:
                                if godziny_abo - tsk_tim_to_add2 >= 0:
                                    godziny_abo -= tsk_tim_to_add2
                                else:
                                    tocount = tsk_tim_to_add2 - godziny_abo
                                    tab_godz[0] += tocount
                                    godziny_abo = 0
                            else:
                                tab_godz[0] += tsk_tim_to_add2


                        taskrow.append("przeskok")
                        break
                    elif x_key+1==l :
                        if (start.hour >= cofigstart and end.hour <24 )or(start.hour >= 0 and end.hour <cofigkoniec ):
                            tsk_tim_to_add = zaoczasfloat(duration.seconds, zao)
                            czasstring = zaoczasstring(duration.seconds, zao)
                            taskrow.append(czasstring)

                            if godziny_abo > 0 and czy_abo == 1 and abologic == 1:
                                if godziny_abo - tsk_tim_to_add >= 0:
                                    godziny_abo -= tsk_tim_to_add
                                else:
                                    tocount = tsk_tim_to_add - godziny_abo
                                    tab_godz[x_key] += tocount
                                    godziny_abo = 0
                            else:
                                tab_godz[x_key] += tsk_tim_to_add

                            taskrow.append("noc")
                        elif (start.hour >= cofigstart and end.hour>=cofigkoniec )or(start.hour >= 0 and end.hour>=cofigkoniec ):
                            taskrow.append("")
                            taskrow.append("")
                            taskrow.append("")
                            taskrow.append("przeskok noc-dzień")
                            firtime = cofigkoniec * 3600 - start.hour * 3600 - start.minute * 60 - start.second

                            czasst = zaoczasstring(firtime, zao)
                            tsk_tim_to_add = zaoczasfloat(firtime, zao)
                            zaoxsec = int(zao * 3600)

                            if przeskoklogicbeta == 1:
                                if zaoxsec < duration.seconds:

                                    taskrow[4]=(czasst)

                                    if godziny_abo > 0 and czy_abo == 1 and abologic == 1:
                                        if godziny_abo - tsk_tim_to_add >= 0:
                                            godziny_abo -= tsk_tim_to_add
                                        else:
                                            tocount = tsk_tim_to_add - godziny_abo
                                            tab_godz[x_key] += tocount
                                            godziny_abo = 0
                                    else:
                                        tab_godz[x_key] += tsk_tim_to_add
                                    sectime = duration.seconds - firtime
                                else:
                                    taskrow.append(" ")

                                    tab_godz[x_key] = 0
                            else:

                                taskrow[4] = (czasst)

                                if godziny_abo > 0 and czy_abo == 1 and abologic == 1:
                                    if godziny_abo - tsk_tim_to_add >= 0:
                                        godziny_abo -= tsk_tim_to_add
                                    else:
                                        tocount = tsk_tim_to_add - godziny_abo
                                        tab_godz[x_key] += tocount
                                        godziny_abo = 0
                                else:
                                    tab_godz[x_key] += tsk_tim_to_add
                            sectime = duration.seconds - firtime


                            nextzao = float(czasy_stawki[0]["zaokraglanie"])
                            czasst = zaoczasstring(sectime, nextzao)
                            taskrow[2] = (czasst)
                            tsk_tim_to_add2 = zaoczasfloat(sectime, nextzao)

                            if godziny_abo > 0 and czy_abo == 1 and abologic == 1:
                                if godziny_abo - tsk_tim_to_add2 >= 0:
                                    godziny_abo -= tsk_tim_to_add2
                                else:
                                    tocount = tsk_tim_to_add2 - godziny_abo
                                    tab_godz[0] += tocount
                                    godziny_abo = 0
                            else:
                                tab_godz[0] += tsk_tim_to_add2



                    else:
                        taskrow.append("-")

                # print(taskrow)
                # print(len(taskrow))

                if len(taskrow) > 0:
                    ws.append(taskrow)
    xls_rest_row.append("Podsumowanie")
    xls_rest_row2.append("")
    xls_rest_row2.append(    project['godziny_abo'] + "*" + project['cenaabo'] + "=" + str(int(project['godziny_abo']) * cenaabo)
)
    # print(tab_godz)
    print("ilosc taskow===" + str(tsccounter))
    # print(tab_godz[0])
    # print(tab_godz[1])
    for r_key, r in tab_godz.items():
        if godziny_abo > 0 and abologic==0:
            if godziny_abo - r >= 0:

                godziny_abo -= r
                tab_godz[r_key] = 0
            else:
                tocount = r - godziny_abo
                tab_godz[r_key] = tocount
                godziny_abo = 0

    xls_rest_row.append(str(int(project['godziny_abo'])-godziny_abo)+"h /"+project['godziny_abo']+"h")
    suma = 0
    for r_key, r in tab_godz.items():
        # print(r_key)
        # print(czasy_stawki[r_key]["stawka"])
        xls_rest_row.append(str(r) + "h")


        #print(r_key)
        if r_key !="Sobota" and r_key !="Niedziela":
            print(r_key)
            suma += r * float(czasy_stawki[r_key]["stawka"])
            xls_rest_row2.append(str(r *float(czasy_stawki[r_key]["stawka"]))+"zł")
            #print(suma)
        elif r_key =="Sobota":
            suma += r * float(weekend["stawkaSobo"])
            xls_rest_row2.append(str(r *float(weekend["stawkaSobo"]))+"zł")
        elif r_key == "Niedziela":
            suma += r * float(weekend["stawkaNied"])
            xls_rest_row2.append(str(r * float(weekend["stawkaNied"])) + "zł")

            #print(suma)

    #print(godziny_abo)
    #print(float(project['godziny_abo'])-godziny_abo)

    suma +=(float(project['godziny_abo']))*cenaabo
    #print(suma)
    xls_rest_row.append("Podsumowanie:" + str(suma) + "zł")
    ws.append(xls_rest_row)
    ws.append(xls_rest_row2)
    greyFill = PatternFill(start_color='F6f6f6',
                          end_color='f6f6f6',
                          fill_type='solid')
    alphlist=string.ascii_uppercase
    column_widths = []
    lastrow =len(ws['A'])
    print(lastrow)
    for g,row in enumerate(ws):
        for k,col in enumerate(row):

            if k==0:
                ws.column_dimensions[get_column_letter(k + 1)].width = 70
            elif k ==l+3:
                 ws.column_dimensions[get_column_letter(k + 1)].width = 30
            else:
                ws.column_dimensions[get_column_letter(k + 1)].width = 20
            if g==0:
                ws[str(alphlist[k])+str(g+1)].fill = greyFill
            if g+1 == lastrow:
                ws[str(alphlist[k])+str(g+1)].fill = greyFill
                ws[str(alphlist[k]) + str(g )].fill = greyFill

    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = column_width+10



    wb.save(pathxls)


